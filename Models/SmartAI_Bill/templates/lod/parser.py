import os
import openpyxl
import csv
import re

ADDRESS_COLUMNS = ["ADDRESS_1", "ADDRESS_2", "ADDRESS_3", "ADDRESS_4", "ADDRESS_5"]


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _format_balance(value):
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return _clean(value)


def parse_lod(file_path: str, limit=None, offset=0) -> dict:
    """
    Parses an LOD Excel (.xlsx) or CSV (.csv) file.
    Handles temporary worker .processing file extensions seamlessly.
    Supports offset and limit for batch slice processing.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"LOD file not found: {file_path}")

    clean_path = file_path[:-11] if file_path.lower().endswith(".processing") else file_path
    ext = os.path.splitext(clean_path)[1].lower()
    all_data = []

    if ext == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                account_no = _clean(row.get("ACCOUNT_NO") or row.get("account_number"))
                if not account_no:
                    continue

                address_lines = [
                    _clean(row.get(col)) for col in ADDRESS_COLUMNS
                    if _clean(row.get(col))
                ]

                all_data.append({
                    "client_name": _clean(row.get("CUSTOMER_NAME") or row.get("client_name")),
                    "client_address_lines": address_lines,
                    "outstanding_balance": _format_balance(row.get("ARREARS") or row.get("outstanding_balance")),
                    "account_number": account_no,
                    "telephone_number": _clean(row.get("EVENT_SOURCE") or row.get("telephone_number")),
                    "regional_office": _clean(row.get("BILLING_CENTRE") or row.get("regional_office")),
                    "reference_number": str(len(all_data) + 1),
                    "letter_date": _clean(row.get("DATE") or row.get("letter_date")),
                })
    else:
        with open(file_path, "rb") as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            total_rows_est = max(0, (ws.max_row or 0) - 1)
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                wb.close()
                return {"records": [], "account_number": "unknown", "total_records": 0, "input_path": file_path}

            columns = {str(name).strip().upper(): idx for idx, name in enumerate(header) if name is not None}
            target_max = (offset + limit) if limit is not None else None

            for row in rows:
                if target_max is not None and len(all_data) >= target_max:
                    break

                acc_idx = columns.get("ACCOUNT_NO")
                if acc_idx is None or acc_idx >= len(row):
                    continue
                account_no = _clean(row[acc_idx])
                if not account_no:
                    continue

                address_lines = []
                for col in ADDRESS_COLUMNS:
                    col_idx = columns.get(col)
                    if col_idx is not None and col_idx < len(row):
                        val = _clean(row[col_idx])
                        if val:
                            address_lines.append(val)

                name_idx = columns.get("CUSTOMER_NAME")
                arr_idx = columns.get("ARREARS")
                tel_idx = columns.get("EVENT_SOURCE")
                reg_idx = columns.get("BILLING_CENTRE")
                date_idx = columns.get("DATE")

                all_data.append({
                    "client_name": _clean(row[name_idx]) if name_idx is not None and name_idx < len(row) else "",
                    "client_address_lines": address_lines,
                    "outstanding_balance": _format_balance(row[arr_idx]) if arr_idx is not None and arr_idx < len(row) else "0.00",
                    "account_number": account_no,
                    "telephone_number": _clean(row[tel_idx]) if tel_idx is not None and tel_idx < len(row) else "",
                    "regional_office": _clean(row[reg_idx]) if reg_idx is not None and reg_idx < len(row) else "",
                    "reference_number": str(len(all_data) + 1),
                    "letter_date": _clean(row[date_idx]) if date_idx is not None and date_idx < len(row) else "",
                })

            wb.close()

    total_records = max(len(all_data), total_rows_est if 'total_rows_est' in locals() else len(all_data))
    sliced_records = all_data[offset : (offset + limit)] if limit is not None else all_data[offset :]
    raw_acc = (sliced_records[0].get("account_number") or sliced_records[0].get("client_name") or "unknown").strip() if sliced_records else "unknown"
    first_acc = re.sub(r'[^A-Za-z0-9_-]+', '_', raw_acc).strip('_')
    if not first_acc:
        first_acc = "unknown"

    return {
        "records": sliced_records,
        "account_number": first_acc,
        "total_records": total_records,
        "input_path": file_path
    }


def load_all_data(limit=None):
    from templates.lod import config
    return parse_lod(config.CLIENTS_XLSX, limit=limit)["records"]
