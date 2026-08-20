"""
parser.py
=========
Reads the customer spreadsheet and converts each raw row into a clean
`Customer` record ready for rendering.

All the "make the messy data presentable" logic lives here:
  - picking a clean name
  - turning ADDR_FULL into tidy address lines (matching the sample layout)
  - resolving the telephone number with a fallback

Run this file directly to sanity-check the first few parsed records:
    python parser.py
"""

from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl

from . import config as C


@dataclass
class Customer:
    name: str
    address_lines: List[str] = field(default_factory=list)
    telephone: str = ""
    raw: dict = field(default_factory=dict)      # original row, for filenames etc.


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------
def _clean(value) -> str:
    """Stringify a cell, trim, and blank out obvious junk tokens."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {t.lower() for t in C.ADDR_JUNK_TOKENS}:
        return ""
    return s


def _pick_name(row: dict) -> str:
    name = _clean(row.get(C.COL_NAME))
    if not name:
        # fallback column is dirty: strip leading commas / collapse spaces
        name = _clean(row.get(C.COL_NAME_FALLBACK)).lstrip(", ").strip()
    return name


def _resolve_phone(row: dict) -> str:
    tel    = _clean(row.get(C.COL_TELEPHONE))
    status = _clean(row.get(C.COL_TEL_STATUS)).upper()
    valid  = tel not in ("", "0") and status in ("OK", "")
    if valid:
        return tel
    if C.TEL_FALLBACK == "account":
        return _clean(row.get(C.COL_ACCOUNT))
    return ""   # "blank"


def _build_address_lines(row: dict) -> List[str]:
    """
    Turn ADDR_FULL into clean address lines matching the printed sample:
        house/building / street / area / city / zipcode

    Strategy: split ADDR_FULL on commas, then trim the tail
    (zip, country, province, district) so only the local address + city
    remain, and re-append the reliable ZIPCODE column as the last line.
    """
    raw_full = row.get(C.COL_ADDR_FULL)
    zipcode  = _clean(row.get(C.COL_ZIPCODE))

    tokens = [t.strip() for t in str(raw_full or "").split(",")]

    # 1) drop a trailing all-digit token (that's the zip, we re-add it cleanly)
    if tokens and tokens[-1].replace(" ", "").isdigit():
        tokens.pop()
    # 2) drop trailing country ("Sri Lanka")
    if tokens and tokens[-1].lower() in C.COUNTRY_TOKENS:
        tokens.pop()
    # 3) drop ONE trailing province if present
    if tokens and tokens[-1].lower() in C.PROVINCES:
        tokens.pop()
    # 4) drop ONE trailing district if present (only one, so a city that
    #    shares a district name survives as the line above it)
    if tokens and tokens[-1].lower() in C.DISTRICTS:
        tokens.pop()

    # 5) clean remaining local lines: drop empties / dashes, dedupe repeats
    lines: List[str] = []
    for t in tokens:
        t = _clean(t)
        if t and (not lines or lines[-1].lower() != t.lower()):
            lines.append(t)

    # 6) append zipcode as the final line
    if zipcode:
        lines.append(zipcode)

    # 7) cap to the configured maximum (name is counted separately)
    return lines[: C.MAX_ADDR_LINES]


# --------------------------------------------------------------------------
# public API
# --------------------------------------------------------------------------
def load_customers(limit: Optional[int] = None) -> List[Customer]:
    return parse_customer_letter(str(C.DATA_XLSX), limit=limit)


def parse_customer_letter(file_path: str, limit: Optional[int] = None, offset: int = 0) -> List[Customer]:
    import os
    import io
    if not os.path.exists(file_path):
        if os.path.exists(str(C.DATA_XLSX)):
            file_path = str(C.DATA_XLSX)
        else:
            raise FileNotFoundError(f"Customer Letter file not found: {file_path}")

    # Safe offset and limit normalization
    try:
        offset_num = int(offset) if offset is not None else 0
    except (ValueError, TypeError):
        offset_num = 0
    if offset_num < 0:
        offset_num = 0

    try:
        limit_num = int(limit) if limit is not None else None
    except (ValueError, TypeError):
        limit_num = None
    if limit_num is not None and limit_num <= 0:
        return []

    clean_path = file_path[:-11] if file_path.lower().endswith(".processing") else file_path
    customers: List[Customer] = []
    current_idx = 0
    target_max = (offset_num + limit_num) if limit_num is not None else None

    with open(file_path, "rb") as f:
        if str(file_path).lower().endswith(".processing") or not str(clean_path).lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            in_mem = io.BytesIO(f.read())
            wb = openpyxl.load_workbook(in_mem, data_only=True)
        else:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

        ws = wb[C.SHEET_NAME] if C.SHEET_NAME and C.SHEET_NAME in wb.sheetnames else wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        header = None
        for values in rows:
            candidate = [str(h).strip() if h is not None else "" for h in values]
            if (
                C.COL_ACCOUNT in candidate 
                or any("ACCOUNT" in str(c).upper() for c in candidate) 
                or "ADDR_FULL" in candidate
                or "SERIAL_NUM" in candidate
                or "CUSTOMER_REF" in candidate
            ):
                header = candidate
                break

        if header is None:
            header = [f"col_{i}" for i in range(50)]

        for values in rows:
            if values is None or all(v is None for v in values):
                continue

            if offset_num > 0 and current_idx < offset_num:
                current_idx += 1
                continue

            if target_max is not None and current_idx >= target_max:
                break

            row = dict(zip(header, values))
            customers.append(
                Customer(
                    name=_pick_name(row),
                    address_lines=_build_address_lines(row),
                    telephone=_resolve_phone(row),
                    raw=row,
                )
            )
            current_idx += 1

        wb.close()

    return customers


if __name__ == "__main__":
    sample = parse_customer_letter(str(C.DATA_XLSX), limit=8)
    print(f"Parsed {len(sample)} sample records:\n")
    for i, c in enumerate(sample, 1):
        print(f"--- Customer {i} ---")
        print("Name :", c.name)
        for ln in c.address_lines:
            print("     |", ln)
        print("Tel  :", c.telephone)
        print()
