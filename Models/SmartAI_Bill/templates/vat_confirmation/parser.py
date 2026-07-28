"""
Parser for VAT Confirmation templates.
Supports CSV (recipients.csv) and Excel (VAT Customer List (002).xlsx).
Handles temporary worker .processing file extensions seamlessly.
Supports offset and limit for partial batch processing.
"""
import os
import csv
import openpyxl
import re


def _clean(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    return "" if val_str in ("-", ".") else val_str


def parse_vat_confirmation(file_path: str, limit=None, offset=0) -> dict:
    """
    Parses VAT Confirmation CSV or Excel file.
    Supports offset and limit for batch slice processing.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"VAT Confirmation file not found: {file_path}")

    clean_path = file_path[:-11] if file_path.lower().endswith(".processing") else file_path
    ext = os.path.splitext(clean_path)[1].lower()
    all_records = []

    if ext == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rec_name = (row.get("recipient_name") or row.get("name") or "").strip()
                ref = (row.get("reference") or row.get("customer_ref") or "").strip()
                vat_no = (row.get("vat_no") or row.get("vat_registration") or "").strip()

                if not ref and not rec_name:
                    continue

                addr_lines = [
                    rec_name,
                    (row.get("address_line1") or "").strip(),
                    (row.get("address_line2") or "").strip(),
                    (row.get("address_line3") or "").strip(),
                    (row.get("address_line4") or "").strip(),
                ]
                clean_lines = [line for line in addr_lines if line and line not in ("-", ".")]

                all_records.append({
                    "recipient_name": rec_name,
                    "reference": ref,
                    "account_number": ref,
                    "vat_no": vat_no,
                    "address_lines": clean_lines,
                })
    else:
        with open(file_path, "rb") as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            total_rows_est = max(0, (ws.max_row or 0) - 1)
            rows = ws.iter_rows(values_only=True)

            try:
                header = next(rows)
            except StopIteration:
                wb.close()
                return {"records": [], "reference": "unknown", "account_number": "unknown", "total_records": 0, "input_path": file_path}

            col_map = {str(name).strip().upper(): idx for idx, name in enumerate(header) if name is not None}
            def _get_col_idx(*names):
                for name in names:
                    if name in col_map:
                        return col_map[name]
                return None

            ref_idx = _get_col_idx("CUSTOMER_REF", "REFERENCE", "ACCOUNT_NO", "ACCOUNT_NUMBER")
            comp_idx = _get_col_idx("COMPANY_NAME")
            name_idx = _get_col_idx("NAME")
            vat_idx = _get_col_idx("VAT_REGISTRATION", "VAT_NO", "VAT_REGISTRATION_NUMBER")
            addr_cols = ["ADDR_LINE_1", "ADDR_LINE_2", "ADDR_LINE_3", "ADDR_LINE_4", "ADDR_LINE_5", "CITY"]

            target_max = (offset + limit) if limit is not None else None

            for row in rows:
                if target_max is not None and len(all_records) >= target_max:
                    break

                ref = _clean(row[ref_idx]) if ref_idx is not None and ref_idx < len(row) else ""
                comp_name = _clean(row[comp_idx]) if comp_idx is not None and comp_idx < len(row) else ""
                indiv_name = _clean(row[name_idx]) if name_idx is not None and name_idx < len(row) else ""

                recipient_name = comp_name or indiv_name
                if not ref and not recipient_name:
                    continue

                vat_no = _clean(row[vat_idx]) if vat_idx is not None and vat_idx < len(row) else ""

                lines = [recipient_name]
                for col in addr_cols:
                    c_idx = col_map.get(col)
                    if c_idx is not None and c_idx < len(row):
                        v = _clean(row[c_idx])
                        if v:
                            lines.append(v)

                all_records.append({
                    "recipient_name": recipient_name,
                    "reference": ref,
                    "account_number": ref,
                    "vat_no": vat_no,
                    "address_lines": lines,
                })

            wb.close()

    total_records = max(len(all_records), total_rows_est if 'total_rows_est' in locals() else len(all_records))
    sliced_records = all_records[offset : (offset + limit)] if limit is not None else all_records[offset :]
    raw_ref = (sliced_records[0].get("reference") or sliced_records[0].get("account_number") or sliced_records[0].get("recipient_name") or "unknown").strip() if sliced_records else "unknown"
    first_ref = re.sub(r'[^A-Za-z0-9_-]+', '_', raw_ref).strip('_')
    if not first_ref:
        first_ref = "unknown"

    return {
        "records": sliced_records,
        "reference": first_ref,
        "account_number": first_ref,
        "total_records": total_records,
        "input_path": file_path
    }


def load_recipients(csv_path):
    return parse_vat_confirmation(csv_path)["records"]
