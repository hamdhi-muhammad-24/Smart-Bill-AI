"""
parser.py  -  read the spreadsheet and build clean FINAL NOTICE records.

Run directly to sanity-check the mapping:
    python parser.py
"""

from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl

from . import config as C


@dataclass
class Customer:
    name: str
    address_lines: List[str]
    fields: dict                       # telephone, account, amount, due_date, contact
    barcode_value: str
    raw: dict = field(default_factory=dict)


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _format_telephone(product_label: str) -> str:
    p = _s(product_label)
    if C.TEL_FORMAT == "lk94" and p:
        return "94" + p.lstrip("0")
    return p


def _format_amount(v) -> str:
    s = _s(v)
    if s == "":
        return ""
    # if already formatted (has a comma), keep as-is
    if "," in s:
        return s
    try:
        return f"{float(s.replace(',', '')):,.{C.AMOUNT_DECIMALS}f}"
    except ValueError:
        return s


def _address_lines(row: dict) -> List[str]:
    lines = [_s(row.get(c)) for c in C.COL_ADDR]
    lines = [ln for ln in lines if ln]          # drop blank address slots
    zc = _s(row.get(C.COL_ZIP))
    if zc:
        lines.append(zc)
    return lines


def _contact(row: dict) -> str:
    if C.CONTACT_SOURCE == "column":
        return _s(row.get(C.COL_CONTACT))
    return C.CONTACT_FIXED


def build_customer(row: dict) -> Customer:
    telephone = _format_telephone(row.get(C.COL_PRODUCT))
    account   = _s(row.get(C.COL_ACCOUNT))
    fields = dict(
        telephone = telephone,
        account   = account,
        amount    = _format_amount(row.get(C.COL_AMOUNT)),
        date      = _s(row.get(C.COL_DATE)),
        due_date  = _s(row.get(C.COL_DUE)),
        contact   = _contact(row),
    )
    barcode_value = account if C.BARCODE_FIELD == "account" else telephone
    return Customer(
        name=_s(row.get(C.COL_NAME)),
        address_lines=_address_lines(row),
        fields=fields,
        barcode_value=barcode_value,
        raw=row,
    )


def load_customers(limit: Optional[int] = None) -> List[Customer]:
    return parse_final_notice(str(C.DATA_XLSX), limit=limit)


def parse_final_notice(file_path: str, limit: Optional[int] = None, offset: int = 0) -> List[Customer]:
    import os
    if not os.path.exists(file_path):
        if os.path.exists(str(C.DATA_XLSX)):
            file_path = str(C.DATA_XLSX)
        else:
            raise FileNotFoundError(f"Final Notice file not found: {file_path}")

    import io
    if str(file_path).lower().endswith(".processing") or not str(file_path).lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        with open(file_path, "rb") as f:
            in_mem = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(in_mem, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    ws = wb[C.SHEET_NAME] if C.SHEET_NAME and C.SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = None
    for values in rows:
        candidate = [_s(h) for h in values]
        if C.COL_ACCOUNT in candidate or any("ACCOUNT" in str(c).upper() for c in candidate):
            header = candidate
            break
    if header is None:
        header = [f"col_{i}" for i in range(50)]

    out: List[Customer] = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        out.append(build_customer(dict(zip(header, values))))
    wb.close()

    if offset > 0:
        out = out[offset:]
    if limit is not None and limit > 0:
        out = out[:limit]
    return out


if __name__ == "__main__":
    for i, c in enumerate(parse_final_notice(str(C.DATA_XLSX), limit=5), 1):
        print(f"--- {i} ---")
        print("name    :", c.name)
        print("address :", c.address_lines)
        print("fields  :", c.fields)
        print("barcode :", c.barcode_value)
        print()
